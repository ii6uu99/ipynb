import openpyxl

file = "file.xlsx"
wb = openpyxl.load_workbook(file)
sheet = wb.active
row = sheet.max_row
col = sheet.max_column

sheet2 = wb.create_sheet(index=2, title="Converted Data")  # 创建一个新的表

row_count = 1
print('started')

for i in range(1, row+1):
    keywords = sheet.cell(i, col).value
    for k in keywords.split(';'):
        for j in range(1, col-1):
            sheet2.cell(row_count, j).value = sheet.cell(i, j).value
        sheet2.cell(row_count, j).value = k
        row_count += 1

wb.save(file)
print('finished')
